# coding=utf-8  #不加入中文编码会出错

# @Time   :2019/6/24 21:54
# @Author :YTL
# @File   :data_process.py

import openpyxl


class DataProcess():
	
	header = ["position_id", "company_name", "position_name", "money", "work_year","education",
			  "job_advantage", "job_description", "job_address", "job_publisher", "company_size"]
	
	def __init__(self, file_name, save_file_name):
		self._file_name = file_name
		self._save_file_name = save_file_name
		self._title = 'zhaopin'
		self.work_book_init()
		pass
	
	def work_book_init(self):
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = self._title
		for col, val in enumerate(self.header, 1):
			ws.cell(1, col, val)
			pass
		wb.save(self._save_file_name)
	
	def read_txt(self):
		res_data = []
		with open(self._file_name, 'r', encoding='utf8') as f:
			res = []
			for line in f.readlines():
				if "=" not in line:
					if "job_description" in line:
						line = line.split("[")[1] #工作职责后面也有：，不能以：分割
						temp = line[0:-2].replace("'", "").replace("\\n", "").strip().split(",")
						desc = ""
						for t in temp:
							if t.strip() is not "":
								desc += t.strip() + '\n'
						res.append(desc)
					elif "job_address" in line:
						line = line.split(":")[1]
						res.append(''.join(line.replace("'", "").strip("\n").split(",")[0:-2]))
					elif "job_publisher" in line:
						line = line.split(":")[1]
						res.append(line.replace("'", "").replace("[", "").replace("]", "").strip("\n"))
					else:
						res.append(line.split(":")[1].strip().strip('\n'))
				else:
					print(res)
					res_data.append(res)
					res = []
		return res_data
	
	def write_data(self):
		res_data = self.read_txt()
		wb = openpyxl.load_workbook(self._save_file_name)
		ws = wb[self._title]  #ws = wb.get_sheet_by_name(self._title)
		for row, data in enumerate(res_data, 2):
			#print(data)
			for col, val in enumerate(data, 1):
				ws.cell(row, col, val)
		wb.save(self._save_file_name)

if __name__ == "__main__":
	file_name = '.\zhaopin.txt'
	save_file_name = "zhaopin.xlsx"
	data_process = DataProcess(file_name, save_file_name)
	data_process.write_data()
	pass